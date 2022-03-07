#!/usr/bin/env python3
# encoding: utf-8

__author__ = "Oliver Schlueter"
__copyright__ = "Copyright 2022, Dell Technologies"
__license__ = "GPL"
__version__ = "1.2.0"
__email__ = "oliver.schlueter@dell.com"
__status__ = "Production"

""""
#########################################################################################################
#
#  DELL EMC ECS Capacity Excel Report 
#
# Permission is hereby granted, free of charge, to any person obtaining a copy of this software 
#  and associated documentation files (the "Software"), to deal in the Software without restriction, 
#  including without limitation the rights to use, copy, modify, merge, publish, distribute, 
#  sublicense, and/or sell copies of the Software, and to permit persons to whom the Software is 
#  furnished to do so, subject to the following conditions:
#  The above copyright notice and this permission notice shall be included in all copies or substantial 
#  portions of the Software.
#  
#  THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR IMPLIED, INCLUDING BUT NOT 
#  LIMITED TO THE WARRANTIES OF MERCHANTABILITY, FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. 
#  IN NO EVENT SHALL THE AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER LIABILITY, 
#  WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM, OUT OF OR IN CONNECTION WITH THE 
#  SOFTWARE OR THE USE OR OTHER DEALINGS IN THE SOFTWARE.
#
###########################################################################################################

#import modules"""
import argparse
import sys
import os
import re
import json
import requests
import urllib3
import datetime
import logging
from openpyxl import Workbook
from openpyxl import load_workbook

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

###########################################
#        VARIABLE
###########################################
max_namespaces = 500
max_buckets = 500


###########################################
#    Methods
###########################################

def escape_ansi(line):
    ansi_escape = re.compile(r'(\x9B|\x1B\[)[0-?]*[ -/]*[@-~]')
    return ansi_escape.sub('', str(line))


def get_argument():
    global hostaddress, user, password, filename, bucket_reporting, DEBUG

    try:

        # Setup argument parser
        parser = argparse.ArgumentParser()
        parser.add_argument('-H', '--hostname',
                            type=str,
                            help='hostname or IP address and Port',
                            required=True)
        parser.add_argument('-u', '--username',
                            type=str,
                            help='username',
                            required=True)
        parser.add_argument('-p', '--password',
                            type=str,
                            help='user password',
                            required=True)
        parser.add_argument('-f', '--filename',
                            type=str,
                            help='Excel Sheet filename',
                            required=True)
        parser.add_argument('-v', '--verbose',
                            action='store_const', const=True,
                            help='verbose logging',
                            required=False)
        parser.add_argument('-b', '--bucket',
                            action='store_const', const=True,
                            help='reporting based on buckets',
                            required=False)

        args = parser.parse_args()

    except KeyboardInterrupt:
        # handle keyboard interrupt #
        return 0

    hostaddress = args.hostname
    user = args.username
    password = args.password
    filename = args.filename
    bucket_reporting = args.bucket
    DEBUG = args.verbose


###########################################
#    CLASS
###########################################

class ecs:
    # This class permit to connect of the ECS's API

    def __init__(self):
        self.user = user
        self.password = password

    def send_request_billing(self):
        # send a request and get the result as dict
        global ecs_results, ecs_capacity_summary
        ecs_results = []
        global ecs_token

        try:
            # try to get token
            url = 'https://' + hostaddress + '/login'
            r = requests.get(url, verify=False, auth=(self.user, self.password))

            # read access token from returned header
            ecs_token = r.headers['X-SDS-AUTH-TOKEN']

            if DEBUG:
                logging.info('Token: ' + ecs_token)

        except Exception as err:
            logging.error('Not able to get token: ' + str(err))
            print(timestamp + ": Not able to get token: " + str(err))
            sys.exit(1)

        if bucket_reporting:
            try:
                # try to get namespaces using token
                url = 'https://' + hostaddress + '/object/namespaces'
                r = requests.get(url, verify=False,
                                 headers={"X-SDS-AUTH-TOKEN": ecs_token, "Accept": "application/json"})

                ecs_namespaces = json.loads(r.content)['namespace']

                count_namespaces = 0

                for namespace in ecs_namespaces:
                    count_namespaces += 1
                    if count_namespaces > max_namespaces:
                        break
                    current_namespace = namespace["name"]
                    if DEBUG:
                        logging.info('Namespace: ' + current_namespace)

                    # try to get buckets using namespaces
                    url = 'https://' + hostaddress + '/object/bucket?namespace=' + current_namespace
                    r = requests.get(url, verify=False,
                                     headers={"X-SDS-AUTH-TOKEN": ecs_token, "Accept": "application/json"})
                    ecs_buckets = json.loads(r.content)['object_bucket']

                    count_buckets = 0
                    for bucket in ecs_buckets:
                        count_buckets += 1
                        if count_buckets > max_buckets:
                            break
                        current_bucket = bucket["name"]
                        if DEBUG:
                            logging.info('Bucket: ' + current_bucket)

                        # try to get capacity data
                        try:
                            url = 'https://' + hostaddress + '/object/billing/buckets/' + current_namespace + '/' + current_bucket + '/info'
                            r = requests.get(url, verify=False,
                                             headers={"X-SDS-AUTH-TOKEN": ecs_token, "Accept": "application/json"})
                            bucket_billing = json.loads(r.content)
                            bucket_total_objects = bucket_billing["total_objects"]
                            bucket_total_size = float(bucket_billing["total_size"])

                        # if not possible set values to zero
                        except:
                            bucket_total_objects = 0
                            bucket_total_size = 0

                        bucket_data = {"name": current_bucket, "namespace": current_namespace,
                                       "total_objects": bucket_total_objects, "total_size": bucket_total_size}
                        ecs_results.append(bucket_data)

            except Exception as err:
                logging.error('Not able to get bucket data: ' + str(err))
                print(timestamp + ": Not able to get bucket data: " + str(err))
                sys.exit(1)
        else:
            try:
                # try to get namespaces using token
                url = 'https://' + hostaddress + '/dashboard/zones/localzone'
                r = requests.get(url, verify=False,
                                 headers={"X-SDS-AUTH-TOKEN": ecs_token, "Accept": "application/json"})

                ecs_dashboard = json.loads(r.content)

                vds_name = ecs_dashboard['name'] + " on " + hostaddress[0:hostaddress.rindex(":")]

                diskSpaceTotalCurrent = round(
                    float(ecs_dashboard['diskSpaceTotalCurrent'][0]['Space']) / 1024 / 1024 / 1024, 2)
                diskSpaceFreeCurrent = round(
                    float(ecs_dashboard['diskSpaceFreeCurrent'][0]['Space']) / 1024 / 1024 / 1024, 2)
                diskSpaceAllocatedCurrent = round(
                    float(ecs_dashboard['diskSpaceAllocatedCurrent'][0]['Space']) / 1024 / 1024 / 1024, 2)
                diskSpaceReservedCurrent = round(
                    float(ecs_dashboard['diskSpaceReservedCurrent'][0]['Space']) / 1024 / 1024 / 1024, 2)

                ecs_capacity_summary = [vds_name, diskSpaceTotalCurrent, diskSpaceFreeCurrent,
                                        diskSpaceAllocatedCurrent, diskSpaceReservedCurrent]
                if DEBUG:
                    logging.info(ecs_capacity_summary)

            except Exception as err:
                logging.error('Not able to get capacity data: ' + str(err))
                print(timestamp + ": Not able to get capacity data: " + str(err))
                sys.exit(1)

    def process_results(self):
        self.send_request_billing()

        # initiate plugin output
        try:
            if DEBUG:
                logging.info("Create Excel Sheet ...")

            # open Sheet if exists otherwise create a new one
            if os.path.isfile(filename):
                workbook = load_workbook(filename)
            else:
                workbook = Workbook()

            sheet = workbook.active

            # reporting on bucket basis
            if bucket_reporting:

                # if there are no columns start with column 3 and create headers
                if sheet.max_column < 3:
                    sheet.cell(1, 1).value = "Namespace"
                    sheet.cell(1, 2).value = "Bucket"
                    new_column = 3
                else:
                    new_column = sheet.max_column + 1

                if DEBUG:
                    logging.info("Add Date Column at column " + str(new_column))

                # insert current timestamp in new column
                sheet.cell(1, new_column).value = datetime.datetime.now().strftime("%d-%b-%Y (%H:%M:%S)")

                for bucket in ecs_results:
                    bucket_index = 0
                    for i in range(1, sheet.max_row + 1):
                        if sheet.cell(i, 1).value == bucket["namespace"] and sheet.cell(i, 2).value == bucket["name"]:
                            bucket_index = i
                            if DEBUG:
                                logging.info("Bucket " + bucket["name"] + " already exists")
                            break

                    # bucket row is already existing in Sheet
                    if bucket_index > 0:
                        sheet.cell(bucket_index, new_column).value = bucket["total_size"]

                    # bucket is not existing in Sheet
                    else:
                        logging.info("New Bucket " + bucket["name"])
                        new_row = sheet.max_row + 1
                        sheet.cell(new_row, 1).value = bucket["namespace"]
                        sheet.cell(new_row, 2).value = bucket["name"]
                        sheet.cell(new_row, new_column).value = bucket["total_size"]

                        if DEBUG:
                            logging.info("New row inserted at row: " + str(new_row) + ", column: " + str(new_column))
            else:

                # if there are no columns start with column 3 and create headers
                if sheet.max_row < 2:
                    sheet.cell(1, 1).value = "VDC Name: " + ecs_capacity_summary[0];
                    sheet.cell(2, 1).value = "Date"
                    sheet.cell(2, 2).value = "Total"
                    sheet.cell(2, 3).value = "Free"
                    sheet.cell(2, 4).value = "Used"
                    sheet.cell(2, 5).value = "Reserved"
                    new_row = 3
                else:
                    new_row = sheet.max_row + 1

                if DEBUG:
                    logging.info("Add Date at row " + str(new_row))

                # insert current timestamp in new column
                sheet.cell(new_row, 1).value = datetime.datetime.now().strftime("%d-%b-%Y (%H:%M:%S)")

                for i in range(1, 5):
                    sheet.cell(new_row, i + 1).value = ecs_capacity_summary[i]

            # save sheet to disk
            workbook.save(filename=filename)

        except Exception as err:
            logging.error('Error while generating result output: ' + str(err))
            print(timestamp + ": Error while generating result output: " + str(err))
            sys.exit(1)


def main():
    # get and test arguments
    print("================================================================")
    print("  Dell EMC ECS Capacity Report as MS Excel Sheet V", __version__)
    print("================================================================")
    get_argument()

    FORMAT = '%(asctime)-15s %(message)s'
    logging.basicConfig(filename='ecs2xls.log', level=logging.INFO, format=FORMAT)

    logging.info('Started')

    # store timestamp
    global timestamp, metric_filter_file, metric_config_file
    timestamp = datetime.datetime.now().strftime("%d-%b-%Y (%H:%M:%S)")

    # display arguments if DEBUG enabled
    if DEBUG:
        logging.info("hostname: " + hostaddress)
        logging.info("user: " + user)
        logging.info("password: " + password)
    else:
        sys.tracebacklimit = 0

    myecs = ecs()
    myecs.process_results()

    logging.info('Finished')


if __name__ == '__main__':
    main()
    sys.exit(0)

